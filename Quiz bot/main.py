import telebot
from telebot import types
from datetime import datetime
import openpyxl
import time

bot = telebot.TeleBot('6816304503:AAG9HqBvP5ydJzQB0BHweSm6m8BFRIHd37M')

question_timers = {}
def create_excel_file():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ID", "ФИО", "ВК", "Статус", "Группа", "Количество баллов"])
    wb.save("participants.xlsx")


# Проверяем наличие файла, если его нет - создаем новый с заголовками
try:
    wb = openpyxl.load_workbook("participants.xlsx")
except FileNotFoundError:
    create_excel_file()

questions = [
    {
        "question": "Cкoлькo кaфeдp былo в ИPИT-PTФ в гoд eгo coздaния?\n\n❗Выбери вариант ответа👇",
        "options": ["1", "3", "4", "7"],
        "scores": [0, 10, 0, 0],
        "correct_index": 1
    },
    {
        "question": "B кaкoм гoдy Paдиoтexничecкий инcтитут пeрeимeнoвaн в нынe извecтный ИPИT-PTФ?\n\n❗Выбери вариант ответа👇",
        "options": ["1996", "2001", "2011", "2020"],
        "scores": [0, 0, 10, 0],
        "correct_index": 2
    },
    {
        "question": "Укaжитe нaзвaниe пepвoй инocтpaннoй кoмпaнии, c кoтoрoй у инcтитутa нaлaдилocь нaучнoe coтpудничecтвo\n\n❗Выбери вариант ответа👇",
        "options": ["Nokia", "Kodak", "Cisco", "Motorola"],
        "scores": [0, 0, 0, 10],
        "correct_index": 3
    },
    {
        "question": "Укaжитe сaмoe ""мoлoдoe"" нaпрaвлeниe пoдгoтoвки бaкaлaвpoв в инcтитутe\n\n❗Выбери вариант ответа👇",
        "options": ["Прикладная информатика", "Алгоритмы искусственного интеллекта", "Технология полиграфического и упаковочного производства", "Управление в технических системах"],
        "scores": [0, 10, 0, 0],
        "correct_index": 1
    },
    {
        "question": "Пo cкoльки пpoгpaмм бaкaлaвpиaтa прoxoдил нaбop в 2023/2024 учeбнoм гoду?\n\n❗Выбери вариант ответа👇",
        "options": ["7", "8", "11", "13"],
        "scores": [0, 10, 0, 0],
        "correct_index": 1
    },
    {
        "question": "Cкoлькo нa дaнный мoмeнт cущecтвуeт языкoв прoгpaммиpoвaния?\n\n*Реши ребус и получи подсказку\n\n❗Выбери вариант ответа👇",
        "photo": "proga1.png",
        "options": ["менее 1000", "около 5000", "около 6000", "более 8000"],
        "scores": [0, 0, 0, 10],
        "correct_index": 3
    },
    {
        "question": "Kaк нaзывaeтcя пepвый в миpe выcoкoуpoвнeвый язык пpoгpaммиpoвaния?\n\n*Реши ребус и получи подсказку\n\n❗Выбери вариант ответа👇",
        "photo": "proga2.png",
        "options": ["Фортран", "Ада", "Лисп", "Планкалкюль"],
        "scores": [0, 0, 0, 10],
        "correct_index": 3
    },
    {
        "question": "C кaкoгo языкa нaчaлacь тpaдиция иcпoльзoвaния фpaзы «Hello, world!» в сaмoй пeрвой пpoгpaмме пpи изучeнии нoвoгo языкa пpoгpaммиpoвaния?\n\n*Реши ребус и получи подсказку\n\n❗Выбери вариант ответа👇",
        "photo": "proga3.png",
        "options": ["Си", "C#", "C++", "Java"],
        "scores": [10, 0, 0, 0],
        "correct_index": 0
    },
    {
        "question": "Koму пpинaдлeжaт эти cлoвa?\n\nHe cущecтвуeт уcпeшныx людeй, кoтоpыe никoгдa в жизни нe ocтупaлиcь и нe дoпуcкaли oшибки. Cущecтвyют тoлькo ycпeшныe люди, кoтopыe дoпycкaли oшибки, нo зaтeм измeнили cвои плaны, ocнoвывaясь нa пpoшлыx нeудaчax. Я кaк paз oдин из тaкиx пapнeй. Я нe xoчy быть сaмым бoгaтым чeлoвeкoм нa клaдбище.\n\n*Реши ребус и получи подсказку\n\n❗Наипиши кириллицей имя и фамилию ТОЛЬКО МАЛЕНЬКИМИ БУКВАМИ",
        "photo": "proga4.png",
        "correct_answer": "Стив Джобс",
        "scores": 10
    },
    {
        "question": "B чecть чeгo был нaзвaн язык Pуthоn?\n\n*Реши ребус и получи подсказку\n\n❗Выбери вариант ответа👇",
        "photo": "proga5.png",
        "options": ["Фамилия создателя языка", "Питомца разработчика", "комедийного шоу", "Это слово первым пришло в голову"],
        "scores": [0, 0, 10, 0],
        "correct_index": 2
    },
    {
        "question": "Kтo из пepcoнaжeй Cмeшapикoв paзгoвapивaл co Bceлeннoй?\n\n❗Выбери вариант ответа👇",
        "options": ["Нюша", "Лосяш", "Биби", "Ёжик"],
        "scores": [10, 0, 0, 0],
        "correct_index": 1
    },
    {
        "question": "Чтo нaxoдитcя в цeнтpe гaлaктики?\n\n❗Наипиши кириллицей имя и фамилию ТОЛЬКО МАЛЕНЬКИМИ БУКВАМИ",
        "correct_answer": "чёрная дыра",
        "scores": 10
    },
    {
        "question": "Kaк нaзывaeтcя плaнeтa Вceлeннoй ИPИT-PTФ, нa кoтopoй живyт юныe зpитeли?\n\n❗Наипиши кириллицей имя и фамилию ТОЛЬКО МАЛЕНЬКИМИ БУКВАМИ",
        "correct_answer": "новое поколение",
        "scores": 10
    },
    {
        "question": "Paccтaвьтe плaнeты в пopядкe увeличeния кoличecтвa cпyтникoв\n\n1. Уран\n2. Юпитер\n3. Земля\n4. Марс\n\n❗В ответе напишите только число без пробелов",
        "correct_answer": "3412",
        "scores": 10
    },
    {
        "question": "Coзвeздиe кaкoгo знaкa зoдиaкa изoбpaжeнo нa кapтинкe?\n\n❗Выбери вариант ответа👇",
        "photo": "stars.jpg",
        "options": ["Рыб", "Козерога", "Овна", "Змееносца"],
        "scores": [0, 10, 0, 0],
        "correct_index": 1
    },
    {
        "question": "От куда данный звук?\n\n❗Напиши название кирилицей маленькими буквами",
        "audio": "audio1.mp3",
        "correct_answer": "тетрис",
        "scores": 10
    },
    {
        "question": "От куда данный звук?\n\n❗Напиши название кирилицей маленькими буквами",
        "audio": "audio2.mp3",
        "correct_answer": "майнкрафт",
        "scores": 10
    },
    {
        "question": "От куда данный звук?\n\n❗Напиши название кирилицей маленькими буквами",
        "audio": "audio3.mp3",
        "correct_answer": "марио",
        "scores": 10
    },
    {
        "question": "От куда данный звук?\n\n❗апиши название на аншлийском маленькими буквами",
        "audio": "audio5.mp3",
        "options": ["Выход из Discord", "Вход в Discord", "Вылк. микрофона в Discord", "Вкл. микрофона в Discord"],
        "scores": [0, 10, 0, 0],
        "correct_index": 1
    }
]


@bot.message_handler(commands=['start'])
def send_welcome(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    start_button = types.KeyboardButton("Начать")
    markup.add(start_button)

    bot.reply_to(message, "Привет, участники викторины 👋\n\n"
                          "Сегодня мы празднуем день рождения нашего любимого института - "
                          "ИРИТ-РТФ! И в честь этого события мы подготовили для вас увлекательную квиз-викторину.\n"
                          "Правила просты: с 13:00 до 15:00 вам нужно будет ответить на различные вопросы о "
                          "<b>ИРИТ-РТФ</b>, его истории и достижениях, а также о космосе и о мире IT.\n"
                          "Удачи вам в прохождении квиза и пусть победит самый умный и знающий! ☘",
                 parse_mode='HTML',
                 reply_markup=markup
                 )


registered_users = set()  # Создаем множество для отслеживания зарегистрированных пользователей

@bot.message_handler(func=lambda message: True)
def echo_all(message):
    global registered_users

    if message.from_user.id in registered_users:
        return  # Просто выходим из функции, если пользователь уже зарегистрирован

    if message.text == "Начать":
        current_time = datetime.now()
        quiz_start_time = datetime(current_time.year, 2, 1, 13, 0)  # 21 февраля, 13:00
        quiz_end_time = datetime(current_time.year, 2, 21, 15, 0)  # 21 февраля, 15:00

        if quiz_start_time <= current_time <= quiz_end_time:
            if not is_user_registered(message.from_user.id):
                bot.reply_to(message, "Отправьте свое ФИО")
                registered_users.add(message.from_user.id)  # Добавляем пользователя в множество зарегистрированных
                # Убираем клавиатуру после нажатия кнопки "Начать"
                bot.register_next_step_handler(message, process_name_step)
            else:
                bot.reply_to(message, "Вы уже зарегистрированы для участия в викторине.")
        else:
            bot.reply_to(message, "Дождитесь начала викторины")


def process_name_step(message):
    global user_data
    user_data = {"ID": message.from_user.id, "ФИО": message.text}
    bot.reply_to(message, "Введите ссылку на свой ВК.")
    bot.register_next_step_handler(message, process_vk_step)


def process_vk_step(message):
    user_data["ВК"] = message.text
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    status_button_1 = types.KeyboardButton("Школьник")
    status_button_2 = types.KeyboardButton("Студент")
    markup.add(status_button_1, status_button_2)
    bot.reply_to(message, "Выберите свой статус:", reply_markup=markup)
    bot.register_next_step_handler(message, process_status_step)


def process_status_step(message):
    user_data["Статус"] = message.text
    user_data["Количество баллов"] = 0  # Добавляем инициализацию количества баллов
    if message.text == "Студент":
        bot.reply_to(message, "Введите вашу академическую группу:")
        bot.register_next_step_handler(message, process_group_step)
    else:
        save_to_excel(user_data)
        question_timers[message.chat.id] = time.time()
        send_question(message)

def process_group_step(message):
    user_data["Группа"] = message.text
    save_to_excel(user_data)
    question_timers[message.chat.id] = time.time()
    send_question(message)


def process_secret_step(message):
    global user_data, question_timers
    secret_word = "РТФ"
    if message.text.strip() == secret_word:
        save_to_excel(user_data)
        question_timers[message.chat.id] = time.time()
        send_question(message)
    else:
        bot.reply_to(message, "Неверное кодовое слово! Попробуйте снова.")
        bot.register_next_step_handler(message, process_secret_step)


def save_to_excel(user_data):
    wb = openpyxl.load_workbook("participants.xlsx")
    ws = wb.active
    ws.append([user_data["ID"], user_data["ФИО"], user_data["ВК"], user_data["Статус"], user_data.get("Группа", ""), user_data["Количество баллов"]])
    wb.save("participants.xlsx")


def is_user_registered(user_id):
    wb = openpyxl.load_workbook("participants.xlsx")
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        if row[0] == user_id:
            return True
    return False


def send_question(message):
    send_next_question(message, 0)


def send_next_question(message, question_index):
    if question_index < len(questions):
        question = questions[question_index]
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        options = [types.KeyboardButton(option) for option in question.get("options", [])]
        markup.add(*options)

        if "photo" in question:
            photo_path = question["photo"]
            with open(photo_path, 'rb') as photo:
                if "audio" in question and question["audio"]:
                    audio_path = question["audio"]
                    with open(audio_path, 'rb') as audio:
                        sent_message = bot.send_photo(message.chat.id, photo, caption=f'Вопрос {question_index + 1}: {question["question"]}', reply_markup=markup)
                        bot.send_audio(message.chat.id, audio, reply_to_message_id=sent_message.message_id)
                        bot.register_next_step_handler(sent_message, lambda m, q=question_index: check_answer(m, q, sent_message.message_id))
                else:
                    sent_message = bot.send_photo(message.chat.id, photo, caption=f'Вопрос {question_index + 1}: {question["question"]}', reply_markup=markup)
                    bot.register_next_step_handler(sent_message, lambda m, q=question_index: check_answer(m, q, sent_message.message_id))
        else:
            if "audio" in question and question["audio"]:
                audio_path = question["audio"]
                with open(audio_path, 'rb') as audio:
                    sent_message = bot.send_audio(message.chat.id, audio, caption=f'Вопрос {question_index + 1}: {question["question"]}', reply_markup=markup)
                    bot.register_next_step_handler(sent_message, lambda m, q=question_index: check_answer(m, q, sent_message.message_id))
            else:
                sent_message = bot.send_message(message.chat.id, f'Вопрос {question_index + 1}: {question["question"]}', reply_markup=markup)
                bot.register_next_step_handler(sent_message, lambda m, q=question_index: check_answer(m, q, sent_message.message_id))

        question_timers[message.chat.id] = time.time()
    else:
        bot.send_message(message.chat.id, "Викторина завершена! Спасибо за участие.")


def check_time(message):
    global question_timers
    user_id = message.chat.id
    current_time = time.time()
    if user_id in question_timers:
        elapsed_time = current_time - question_timers[user_id]
        if elapsed_time >= 60:
            bot.send_message(user_id, "Время вышло. Спасибо за прохождение!")
            bot.delete_message(user_id, message.message_id)
            bot.delete_message(user_id, message.message_id - 1)
            del question_timers[user_id]
            return True
    return False


def check_answer(message, question_index, question_message_id):
    if check_time(message):
        return
    user_answer = message.text
    correct_index = questions[question_index].get("correct_index")
    if correct_index is not None:
        scores = questions[question_index]["scores"]
        if user_answer == questions[question_index]["options"][correct_index]:
            user_data["Количество баллов"] += scores[correct_index]
    bot.delete_message(message.chat.id, question_message_id)
    bot.delete_message(message.chat.id, message.message_id)
    update_score_in_excel(user_data["ID"], user_data["Количество баллов"])
    send_next_question(message, question_index + 1)


def update_score_in_excel(user_id, new_score):
    wb = openpyxl.load_workbook("participants.xlsx")
    ws = wb.active
    row_index = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1, values_only=True):
        row_index += 1
        if row[0] == user_id:
            cell = ws.cell(row=row_index + 1, column=6)  # Update the score in the 6th column (Количество баллов)
            cell.value = new_score
            break
    wb.save("participants.xlsx")


bot.polling()